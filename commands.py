import datetime

import openpyxl
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

from json_compare_common_methods import *
import sys
import json
import os

script_directory = os.path.dirname(os.path.abspath(__file__))


def load_file(file_path):
    with open(file_path, "r") as file:
        file_content = file.read()
    return file_content


def save_json_file(file_path, json_dict):
    with open(file_path, 'w') as json_file:
        json.dump(json_dict, json_file, indent=4)


class Commands:
    HELP = "help"
    CREATE_MAPPING_TEMPLATE = "create-mapping-template"
    CREATE_SORTING_RULES = "create-sorting-rules"
    COMPARE = "compare"


command = Commands()


class Command:

    def run(self, args):
        if len(args) < 2:
            print(f"Missing arguments, do --help to check usage")
        if args[1][2:].startswith(command.HELP):
            HelpCmd().run(args)
        if args[1][2:].startswith(command.CREATE_MAPPING_TEMPLATE):
            MappingTemplateCmd().run(args[2:])
        if args[1][2:].startswith(command.COMPARE):
            CompareCmd().run(args[2:])


class HelpCmd(Commands):
    def run(self, args):
        print(f"\nUsage: "
              f"\npy json-compare-tools.py --option <arguments>"
              f"\npy json-compare-tools.py --{command.HELP} show usage"
              f"\npy json-compare-tools.py --{command.CREATE_MAPPING_TEMPLATE} exp_json_file dps_json_file mapping_template_file"
              # f"\npy json-compare-tools.py --{command.CREATE_SORTING_RULES} mapping_template_file"
              f"\npy json-compare-tools.py --{command.COMPARE} exp_json_file dps_json_file mapping_template_file sorting_rules_file")


class MappingTemplateCmd(Commands):

    def run(self, args):
        if len(args) >= 2:
            exp_json_file = args[0]
            dps_json_file = args[1]
            mapping_template_file = list_get(args, 2, None) if not is_blank(list_get(args, 2, None)) else None
            self.__process(exp_json_file, dps_json_file, mapping_template_file)
        else:
            print(f"{command.CREATE_MAPPING_TEMPLATE} require at least 2 arguments but only have {len(args)}")

    def __process(self, exp_json_file, dps_json_file, mapping_template_file=None):
        exp_json = load_file(exp_json_file)
        dps_json = load_file(dps_json_file)
        mapping_template = {}
        if not is_blank(mapping_template_file):
            mapping_template = load_file(mapping_template_file)
        schema_exp = json_to_schema(json.loads(exp_json))
        schema_dps = json_to_schema(json.loads(dps_json))
        json_exp_paths = list_all_paths(schema_exp)
        json_dps_paths = list_all_paths(schema_dps)
        auto_mapping_template = auto_mapping(json_exp_paths, json_dps_paths, mapping_template)
        try:
            final_mapping_template = create_mapping_template(json_exp_paths, json_dps_paths, auto_mapping_template)
        except:
            final_mapping_template = auto_mapping_template
            print("No Console Screen skip manually create mapping template")
        self.__save_template(final_mapping_template)

    def __save_template(self, mapping_template):
        print("Enter a name to save mapping template")
        file_name = input()
        folder_path = os.path.join(script_directory, 'save', 'mapping_template')
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        file_path = f"{folder_path}\\{file_name}.json"
        save_json_file(file_path, mapping_template)
        print(f"Template save in {file_path}")


class CompareCmd(Commands):

    def run(self, args):
        if len(args) >= 3:
            exp_json_file = list_get(args, 0, None)
            dps_json_file = list_get(args, 1, None)
            mapping_template_file = list_get(args, 2, None)
            self.__process(exp_json_file, dps_json_file, mapping_template_file)
        else:
            print(f"{command.CREATE_MAPPING_TEMPLATE} require at least 2 arguments but only have {len(args)}")

    def __process(self, exp_json_file, dps_json_file, mapping_template_file=None):
        exp_json = load_file(exp_json_file)
        dps_json = load_file(dps_json_file)
        mapping_template = load_file(mapping_template_file)
        dps_stdzng_path_value_dict_out = map_b_paths_to_a(json.loads(mapping_template), extract_paths_values(json.loads(dps_json)))
        result = compare_dicts(extract_paths_values(json.loads(exp_json)), extract_paths_values(dps_stdzng_path_value_dict_out))
        self.__save_result(result)

    def __save_result(self, result):
        today = datetime.date.today()
        formatted_date = today.strftime("%Y%m%d")
        folder_path = os.path.join(script_directory, 'result', formatted_date)
        current_datetime = datetime.datetime.now()
        formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S_%f")[:-3]
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        json_file_path = f"{folder_path}\\result_{formatted_datetime}.json"
        save_json_file(json_file_path, result)
        excel_file_path = f"{folder_path}\\result_{formatted_datetime}.xlsx"
        self.save_excel_file(excel_file_path, result)
        print(f"result json in {json_file_path}", result)
        print(f"result execel in {excel_file_path}")

    def save_excel_file(self, excel_file_path, result):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet.append(["Path", "EXP Value", "DPS Value", "Is Match"])

        fill_not_match = PatternFill(start_color="FFCCFF", end_color="FFCCFF", fill_type="solid")

        for diff_item in result:
            row = [diff_item["path"], diff_item["exp_value"], diff_item["dps_value"], diff_item["isMatch"]]
            sheet.append(row)

            if not diff_item["isMatch"]:
                sheet[sheet.max_row][-1].fill = fill_not_match

            for col_idx in [2, 3]:
                cell = sheet.cell(row=sheet.max_row, column=col_idx)
                cell.alignment = Alignment(horizontal="right")

        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        workbook.save(excel_file_path)